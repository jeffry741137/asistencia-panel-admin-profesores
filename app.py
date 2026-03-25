import os, calendar, json, io, base64, unicodedata, secrets, hashlib
from datetime import datetime, timezone, timedelta
from flask import Flask, request, jsonify, send_file, Response, session
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import psycopg2
from psycopg2.extras import RealDictCursor
from psycopg2 import pool as pg_pool

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = False
CORS(app, supports_credentials=True)

# ── Neon PostgreSQL ────────────────────────────────────────────────────────────
DATABASE_URL = os.environ.get("postgresql://neondb_owner:npg_4xWl7ZFLyzum@ep-autumn-paper-a4oixcp9-pooler.us-east-1.aws.neon.tech/neondb?sslmode=require&channel_binding=require", "")
_pool = None

def get_pool():
    global _pool
    if _pool is None:
        _pool = pg_pool.SimpleConnectionPool(1, 10, DATABASE_URL, sslmode="require")
    return _pool

def db_exec(sql, params=(), fetch=None):
    conn = get_pool().getconn()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(sql, params)
            conn.commit()
            if fetch == "one":
                row = cur.fetchone()
                return dict(row) if row else None
            if fetch == "all":
                return [dict(r) for r in cur.fetchall()]
            return None
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        get_pool().putconn(conn)

# ── Constantes ─────────────────────────────────────────────────────────────────
CREATOR_USER = "JeffryPeHu"
CREATOR_PASS_HASH = hashlib.sha256(
    os.environ.get("CREATOR_PASS", "Creator2026!#").encode()
).hexdigest()

MESES_ES = {1:"ENERO",2:"FEBRERO",3:"MARZO",4:"ABRIL",5:"MAYO",6:"JUNIO",
            7:"JULIO",8:"AGOSTO",9:"SEPTIEMBRE",10:"OCTUBRE",11:"NOVIEMBRE",12:"DICIEMBRE"}
PERU_TZ = timezone(timedelta(hours=-5))
HORA_LIMITE = (7, 55)
HORA_FALTA  = (16, 0)

# ── Init DB ────────────────────────────────────────────────────────────────────
def init_db():
    db_exec("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            nombre_completo TEXT DEFAULT '',
            institucion TEXT DEFAULT '',
            salon TEXT DEFAULT '',
            activo BOOLEAN DEFAULT TRUE,
            creado_en TIMESTAMPTZ DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS codigos_invitacion (
            id SERIAL PRIMARY KEY,
            codigo TEXT UNIQUE NOT NULL,
            usado BOOLEAN DEFAULT FALSE,
            creado_en TIMESTAMPTZ DEFAULT NOW(),
            expira_en TIMESTAMPTZ NOT NULL,
            usado_por TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS alumnos (
            id SERIAL PRIMARY KEY,
            usuario_id INTEGER REFERENCES usuarios(id) ON DELETE CASCADE,
            nombre TEXT NOT NULL,
            genero TEXT CHECK(genero IN ('nino','nina')) DEFAULT 'nino',
            orden INTEGER DEFAULT 0,
            UNIQUE(usuario_id, nombre)
        );
        CREATE TABLE IF NOT EXISTS asistencia (
            id SERIAL PRIMARY KEY,
            usuario_id INTEGER REFERENCES usuarios(id) ON DELETE CASCADE,
            clave TEXT NOT NULL,
            alumno TEXT NOT NULL,
            dia TEXT NOT NULL,
            marca TEXT NOT NULL,
            hora TEXT NOT NULL,
            motivo TEXT DEFAULT '',
            UNIQUE(usuario_id, clave, alumno, dia)
        );
    """)

# ── Auth helpers ───────────────────────────────────────────────────────────────
def hash_pass(p):
    return hashlib.sha256(p.encode()).hexdigest()

def get_current_user():
    if session.get("is_creator"):
        return {"id": 0, "username": CREATOR_USER, "is_creator": True,
                "activo": True, "salon": "ADMIN", "institucion": "Sistema"}
    uid = session.get("user_id")
    if not uid:
        return None
    u = db_exec("SELECT * FROM usuarios WHERE id=%s", (uid,), fetch="one")
    if not u:
        return None
    u["is_creator"] = False
    return u

def require_login():
    u = get_current_user()
    if not u:
        return jsonify({"ok": False, "msg": "no_auth"}), 401
    if not u.get("activo") and not u.get("is_creator"):
        return jsonify({"ok": False, "msg": "cuenta_desactivada"}), 403
    return u

# ── Tiempo Perú ────────────────────────────────────────────────────────────────
def ahora_peru():
    return datetime.now(timezone.utc).astimezone(PERU_TZ)

def info_hoy():
    n = ahora_peru()
    return {"dia": n.day, "mes": n.month, "ano": n.year,
            "mes_nom": MESES_ES[n.month], "hora": n.strftime("%H:%M"),
            "hora_h": n.hour, "hora_m": n.minute}

def dias_habiles(anio, mes):
    _, total = calendar.monthrange(anio, mes)
    return [d for d in range(1, total+1) if calendar.weekday(anio, mes, d) < 5]

def dia_letra(anio, mes, dia):
    return ["L","M","M","J","V"][calendar.weekday(anio, mes, dia)]

def es_temprano(h, m):
    return h*60+m <= HORA_LIMITE[0]*60+HORA_LIMITE[1]

def es_hora_falta():
    n = ahora_peru()
    return n.hour*60+n.minute >= HORA_FALTA[0]*60+HORA_FALTA[1]

def normalizar(s):
    return unicodedata.normalize("NFD", s.lower()).encode("ascii","ignore").decode()

def buscar_alumno_en_lista(lista, query_str):
    q = normalizar(query_str.strip())
    if not q:
        return None
    for a in lista:
        if normalizar(a) == q:
            return a
    matches = [a for a in lista if q in normalizar(a)]
    return matches[0] if matches else None

# ── DB asistencia ──────────────────────────────────────────────────────────────
def db_get_asistencia(uid, clave=None, alumno=None, dia=None):
    sql = "SELECT * FROM asistencia WHERE usuario_id=%s"
    params = [uid]
    if clave:  sql += " AND clave=%s";  params.append(clave)
    if alumno: sql += " AND alumno=%s"; params.append(alumno)
    if dia:    sql += " AND dia=%s";    params.append(dia)
    return db_exec(sql, params, fetch="all") or []

def db_upsert_asistencia(uid, clave, alumno, dia, marca, hora, motivo=""):
    db_exec("""
        INSERT INTO asistencia (usuario_id,clave,alumno,dia,marca,hora,motivo)
        VALUES (%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (usuario_id,clave,alumno,dia)
        DO UPDATE SET marca=EXCLUDED.marca, hora=EXCLUDED.hora, motivo=EXCLUDED.motivo
    """, (uid, clave, alumno, dia, marca, hora, motivo))

def get_alumnos_usuario(uid):
    rows = db_exec(
        "SELECT nombre, genero FROM alumnos WHERE usuario_id=%s ORDER BY orden, nombre",
        (uid,), fetch="all"
    ) or []
    return rows

def marcar_faltas_usuario(uid):
    if not es_hora_falta():
        return 0
    hoy = info_hoy()
    clave = f"{hoy['ano']}-{hoy['mes']:02d}"
    dia = str(hoy["dia"])
    alumnos = [r["nombre"] for r in get_alumnos_usuario(uid)]
    regs_dia = db_get_asistencia(uid, clave=clave, dia=dia)
    con_reg = {r["alumno"] for r in regs_dia}
    count = 0
    for alumno in alumnos:
        if alumno not in con_reg:
            db_upsert_asistencia(uid, clave, alumno, dia, "F", "--:--")
            count += 1
    return count

# ── Excel ──────────────────────────────────────────────────────────────────────
LOGO_B64 = ""  # se puede poner logo en base64 si se desea

def generar_excel(uid, anio, mes, salon=""):
    dias = dias_habiles(anio, mes)
    mes_nom = MESES_ES[mes]
    clave = f"{anio}-{mes:02d}"
    alumnos_rows = get_alumnos_usuario(uid)
    ALUMNOS = [r["nombre"] for r in alumnos_rows]
    if not ALUMNOS:
        ALUMNOS = []

    regs = db_get_asistencia(uid, clave=clave)
    data = {}
    for r in regs:
        a = r["alumno"]
        if a not in data:
            data[a] = {}
        data[a][r["dia"]] = {"marca": r["marca"], "hora": r["hora"]}

    grupo1 = ALUMNOS[:15]
    grupo2 = ALUMNOS[15:]
    COL_NUM=2; COL_NOM=3; COL_D1=4
    THIN=Side(style="thin",color="BDBDBD")
    MEDIUM=Side(style="medium",color="888888")

    wb = Workbook()
    ws = wb.active
    ws.title = f"{mes_nom} {anio}"
    ws.column_dimensions["A"].width = 1.66
    ws.column_dimensions["B"].width = 2.83
    ws.column_dimensions["C"].width = 35.5
    for i in range(len(dias)):
        ws.column_dimensions[get_column_letter(COL_D1+i)].width = 4.0

    alturas = {1:5.25,2:32,3:35,4:12,5:23,6:3,7:11.25,8:17,9:19}
    for r,h in alturas.items():
        ws.row_dimensions[r].height = h
    for i in range(max(len(grupo1),1)):
        ws.row_dimensions[10+i].height = 23
    ws.row_dimensions[25].height = 14.25
    ws.row_dimensions[27].height = 16
    ws.row_dimensions[28].height = 17
    ws.row_dimensions[29].height = 20
    ws.row_dimensions[30].height = 20
    for i in range(max(len(grupo2),1)):
        ws.row_dimensions[31+i].height = 20

    last_col = COL_D1+len(dias)-1

    ws.merge_cells("D2:T2")
    c = ws["D2"]
    c.value = salon or "ASISTENCIA"
    c.font = Font(bold=True, size=20)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("W2:Z3")
    c = ws["W2"]
    c.value = salon
    c.font = Font(bold=True, size=48)
    c.fill = PatternFill("solid", fgColor="FFFF00")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

    ws.merge_cells("F3:R3")
    c = ws["F3"]
    c.value = f"ASISTENCIA - {anio}"
    c.font = Font(bold=True, size=20)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("W4:Z5")
    c = ws["W4"]
    c.value = mes_nom
    c.font = Font(size=16)
    c.fill = PatternFill("solid", fgColor="CC99FF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

    ws.merge_cells(f"C7:{get_column_letter(last_col)}7")
    ws["C7"].fill = PatternFill("solid", fgColor="D9D9D9")

    def bloque(fc, fd, fi, grupo, off):
        ws.merge_cells(f"B{fc}:C{fd}")
        c = ws.cell(fc, 2, "NOMBRES Y APELLIDOS")
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)
        for i, d in enumerate(dias):
            c = ws.cell(fc, COL_D1+i, dia_letra(anio, mes, d))
            c.font = Font(size=14)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        for i, d in enumerate(dias):
            c = ws.cell(fd, COL_D1+i, d)
            c.font = Font(size=14)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        for idx, nombre in enumerate(grupo):
            fila = fi+idx; num = off+idx+1
            par = idx % 2 == 0; bg = "FFFFFF" if par else "F2F2F2"
            c = ws.cell(fila, COL_NUM, num)
            c.font = Font(size=12)
            c.alignment = Alignment(vertical="center")
            c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            c.fill = PatternFill("solid", fgColor=bg)
            c = ws.cell(fila, COL_NOM, nombre)
            c.font = Font(size=14, color="000000")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = Border(left=MEDIUM, right=MEDIUM, top=THIN, bottom=THIN)
            c.fill = PatternFill("solid", fgColor=bg)
            ar = data.get(nombre, {})
            for i, d in enumerate(dias):
                cell = ws.cell(fila, COL_D1+i)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
                reg = ar.get(str(d))
                if reg:
                    mk = reg["marca"]
                    if mk=="A":
                        cell.value="A"; cell.font=Font(bold=True,size=12,color="1B5E20")
                        cell.fill=PatternFill("solid",fgColor="C8E6C9")
                    elif mk=="T":
                        cell.value="T"; cell.font=Font(bold=True,size=12,color="BF360C")
                        cell.fill=PatternFill("solid",fgColor="FFE0B2")
                    elif mk=="F":
                        cell.value="F"; cell.font=Font(bold=True,size=12,color="FFFFFF")
                        cell.fill=PatternFill("solid",fgColor="C62828")
                    elif mk=="J":
                        cell.value="J"; cell.font=Font(bold=True,size=12,color="FFFFFF")
                        cell.fill=PatternFill("solid",fgColor="1565C0")
                else:
                    cell.fill = PatternFill("solid", fgColor=bg)

    if grupo1:
        bloque(8, 9, 10, grupo1, 0)
    if grupo2:
        ws.merge_cells(f"C27:{get_column_letter(last_col)}27")
        ws["C27"].fill = PatternFill("solid", fgColor="D9D9D9")
        bloque(29, 30, 31, grupo2, 15)

    ws.freeze_panes = "D10"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ══════════════════════════════════════════════════════════════════════════════
# RUTAS AUTH
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/auth/login", methods=["POST"])
def auth_login():
    data = request.get_json(force=True)
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    if not username or not password:
        return jsonify({"ok": False, "msg": "Completa todos los campos"})

    # Cuenta creador
    if username == CREATOR_USER:
        if hash_pass(password) == CREATOR_PASS_HASH:
            session.clear()
            session["is_creator"] = True
            session["username"] = CREATOR_USER
            return jsonify({"ok": True, "is_creator": True, "username": CREATOR_USER})
        return jsonify({"ok": False, "msg": "Contraseña incorrecta"})

    # Usuario normal
    u = db_exec("SELECT * FROM usuarios WHERE username=%s", (username,), fetch="one")
    if not u:
        return jsonify({"ok": False, "msg": "Usuario no encontrado"})
    if not u["activo"]:
        return jsonify({"ok": False, "msg": "Cuenta desactivada. Contacta al administrador"})
    if hash_pass(password) != u["password_hash"]:
        return jsonify({"ok": False, "msg": "Contraseña incorrecta"})

    session.clear()
    session["user_id"] = u["id"]
    session["username"] = u["username"]
    return jsonify({"ok": True, "is_creator": False, "username": u["username"],
                    "salon": u["salon"], "institucion": u["institucion"]})

@app.route("/auth/registro", methods=["POST"])
def auth_registro():
    data = request.get_json(force=True)
    codigo = (data.get("codigo") or "").strip().upper()
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    nombre_completo = (data.get("nombre_completo") or "").strip()
    institucion = (data.get("institucion") or "").strip()
    salon = (data.get("salon") or "").strip()

    if not all([codigo, username, password, nombre_completo, salon]):
        return jsonify({"ok": False, "msg": "Completa todos los campos"})
    if len(password) < 6:
        return jsonify({"ok": False, "msg": "La contraseña debe tener al menos 6 caracteres"})

    # Verificar código
    now_utc = datetime.now(timezone.utc)
    inv = db_exec(
        "SELECT * FROM codigos_invitacion WHERE codigo=%s AND usado=FALSE",
        (codigo,), fetch="one"
    )
    if not inv:
        return jsonify({"ok": False, "msg": "Código inválido o ya usado"})

    expira = inv["expira_en"]
    if expira.tzinfo is None:
        expira = expira.replace(tzinfo=timezone.utc)
    if now_utc > expira:
        return jsonify({"ok": False, "msg": "El código expiró (válido 10 minutos)"})

    # Verificar username único
    existe = db_exec("SELECT id FROM usuarios WHERE username=%s", (username,), fetch="one")
    if existe:
        return jsonify({"ok": False, "msg": "Ese nombre de usuario ya está en uso"})

    # Crear usuario
    u = db_exec("""
        INSERT INTO usuarios (username, password_hash, nombre_completo, institucion, salon)
        VALUES (%s,%s,%s,%s,%s) RETURNING id
    """, (username, hash_pass(password), nombre_completo, institucion, salon), fetch="one")

    # Marcar código como usado
    db_exec(
        "UPDATE codigos_invitacion SET usado=TRUE, usado_por=%s WHERE id=%s",
        (username, inv["id"])
    )

    session.clear()
    session["user_id"] = u["id"]
    session["username"] = username
    return jsonify({"ok": True, "username": username, "salon": salon})

@app.route("/auth/logout", methods=["POST"])
def auth_logout():
    session.clear()
    return jsonify({"ok": True})

@app.route("/auth/me")
def auth_me():
    u = get_current_user()
    if not u:
        return jsonify({"ok": False, "msg": "no_auth"})
    return jsonify({"ok": True, "is_creator": u.get("is_creator", False),
                    "username": u.get("username"), "salon": u.get("salon",""),
                    "institucion": u.get("institucion",""), "activo": u.get("activo", True)})

# ══════════════════════════════════════════════════════════════════════════════
# RUTAS ADMIN (solo CREADOR)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/generar_codigo", methods=["POST"])
def admin_generar_codigo():
    u = get_current_user()
    if not u or not u.get("is_creator"):
        return jsonify({"ok": False, "msg": "Solo el creador puede hacer esto"}), 403

    codigo = secrets.token_urlsafe(8).upper()[:10]
    expira = datetime.now(timezone.utc) + timedelta(minutes=10)
    db_exec(
        "INSERT INTO codigos_invitacion (codigo, expira_en) VALUES (%s,%s)",
        (codigo, expira)
    )
    return jsonify({"ok": True, "codigo": codigo,
                    "expira_en": expira.strftime("%H:%M:%S UTC"),
                    "mensaje": "Válido por 10 minutos, uso único"})

@app.route("/admin/usuarios")
def admin_usuarios():
    u = get_current_user()
    if not u or not u.get("is_creator"):
        return jsonify({"ok": False, "msg": "No autorizado"}), 403

    usuarios = db_exec("""
        SELECT id, username, nombre_completo, institucion, salon, activo,
               creado_en,
               (SELECT COUNT(*) FROM alumnos WHERE usuario_id=u.id) AS num_alumnos
        FROM usuarios u ORDER BY creado_en DESC
    """, fetch="all") or []

    for usr in usuarios:
        if usr.get("creado_en"):
            usr["creado_en"] = usr["creado_en"].strftime("%d/%m/%Y %H:%M")
    return jsonify({"ok": True, "usuarios": usuarios})

@app.route("/admin/toggle_usuario", methods=["POST"])
def admin_toggle_usuario():
    u = get_current_user()
    if not u or not u.get("is_creator"):
        return jsonify({"ok": False, "msg": "No autorizado"}), 403

    data = request.get_json(force=True)
    uid = data.get("id")
    activo = data.get("activo")
    if uid is None or activo is None:
        return jsonify({"ok": False, "msg": "Datos incompletos"})

    db_exec("UPDATE usuarios SET activo=%s WHERE id=%s", (activo, uid))
    estado = "activada" if activo else "desactivada"
    return jsonify({"ok": True, "msg": f"Cuenta {estado} correctamente"})

@app.route("/admin/eliminar_usuario", methods=["POST"])
def admin_eliminar_usuario():
    u = get_current_user()
    if not u or not u.get("is_creator"):
        return jsonify({"ok": False, "msg": "No autorizado"}), 403

    data = request.get_json(force=True)
    uid = data.get("id")
    if not uid:
        return jsonify({"ok": False, "msg": "ID requerido"})

    db_exec("DELETE FROM usuarios WHERE id=%s", (uid,))
    return jsonify({"ok": True, "msg": "Usuario eliminado"})

@app.route("/admin/codigos")
def admin_codigos():
    u = get_current_user()
    if not u or not u.get("is_creator"):
        return jsonify({"ok": False, "msg": "No autorizado"}), 403

    codigos = db_exec("""
        SELECT codigo, usado, usado_por, creado_en, expira_en
        FROM codigos_invitacion ORDER BY creado_en DESC LIMIT 20
    """, fetch="all") or []

    now_utc = datetime.now(timezone.utc)
    for c in codigos:
        expira = c["expira_en"]
        if expira.tzinfo is None:
            expira = expira.replace(tzinfo=timezone.utc)
        c["expirado"] = now_utc > expira
        c["creado_en"] = c["creado_en"].strftime("%d/%m/%Y %H:%M")
        c["expira_en"] = expira.strftime("%d/%m/%Y %H:%M")
    return jsonify({"ok": True, "codigos": codigos})

# ══════════════════════════════════════════════════════════════════════════════
# RUTAS ALUMNOS (gestión por usuario)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/alumnos/lista")
def alumnos_lista():
    result = require_login()
    if isinstance(result, tuple): return result
    u = result
    if u.get("is_creator"):
        return jsonify({"ok": True, "alumnos": []})
    rows = get_alumnos_usuario(u["id"])
    return jsonify({"ok": True, "alumnos": rows})

@app.route("/alumnos/agregar", methods=["POST"])
def alumnos_agregar():
    result = require_login()
    if isinstance(result, tuple): return result
    u = result
    if u.get("is_creator"):
        return jsonify({"ok": False, "msg": "El creador no gestiona alumnos"})

    data = request.get_json(force=True)
    nombre = (data.get("nombre") or "").strip()
    genero = data.get("genero", "nino")
    if not nombre:
        return jsonify({"ok": False, "msg": "Nombre requerido"})
    if genero not in ("nino", "nina"):
        genero = "nino"

    max_orden = db_exec(
        "SELECT COALESCE(MAX(orden),0) AS m FROM alumnos WHERE usuario_id=%s",
        (u["id"],), fetch="one"
    )
    orden = (max_orden["m"] if max_orden else 0) + 1

    try:
        db_exec(
            "INSERT INTO alumnos (usuario_id, nombre, genero, orden) VALUES (%s,%s,%s,%s)",
            (u["id"], nombre, genero, orden)
        )
        return jsonify({"ok": True, "msg": "Alumno agregado"})
    except Exception:
        return jsonify({"ok": False, "msg": "Ese alumno ya existe"})

@app.route("/alumnos/eliminar", methods=["POST"])
def alumnos_eliminar():
    result = require_login()
    if isinstance(result, tuple): return result
    u = result
    data = request.get_json(force=True)
    nombre = (data.get("nombre") or "").strip()
    db_exec("DELETE FROM alumnos WHERE usuario_id=%s AND nombre=%s", (u["id"], nombre))
    return jsonify({"ok": True})

# ══════════════════════════════════════════════════════════════════════════════
# RUTAS ASISTENCIA
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/registrar", methods=["POST"])
def registrar():
    result = require_login()
    if isinstance(result, tuple): return result
    u = result

    req = request.get_json(force=True)
    nombre_query = (req.get("nombre") or "").strip()
    if not nombre_query:
        return jsonify({"ok": False, "msg": "Nombre vacío"})

    alumnos = [r["nombre"] for r in get_alumnos_usuario(u["id"])]
    nombre_real = buscar_alumno_en_lista(alumnos, nombre_query)
    if not nombre_real:
        return jsonify({"ok": False, "msg": f"Alumno no encontrado: {nombre_query}"})

    hoy = info_hoy()
    clave = f"{hoy['ano']}-{hoy['mes']:02d}"
    regs = db_get_asistencia(u["id"], clave=clave, alumno=nombre_real, dia=str(hoy["dia"]))
    if regs and regs[0].get("marca") in ["A","T"]:
        return jsonify({"ok": False, "msg": "ya_registrado", "nombre": nombre_real})

    tipo = "temprano" if es_temprano(hoy["hora_h"], hoy["hora_m"]) else "tardanza"
    marca = "A" if tipo == "temprano" else "T"
    db_upsert_asistencia(u["id"], clave, nombre_real, str(hoy["dia"]), marca, hoy["hora"])

    regs_dia = db_get_asistencia(u["id"], clave=clave, dia=str(hoy["dia"]))
    presente = sum(1 for r in regs_dia if r["marca"] in ["A","T","J"])
    return jsonify({"ok": True, "nombre": nombre_real, "dia": hoy["dia"],
                    "hora": hoy["hora"], "tipo": tipo, "marca": marca,
                    "presente_hoy": presente})

@app.route("/justificar", methods=["POST"])
def justificar():
    result = require_login()
    if isinstance(result, tuple): return result
    u = result

    req = request.get_json(force=True)
    nombre = (req.get("nombre") or "").strip()
    motivo = (req.get("motivo") or "").strip()
    if not nombre: return jsonify({"ok": False, "msg": "Nombre vacío"})
    if not motivo: return jsonify({"ok": False, "msg": "Escribe el motivo"})

    alumnos = [r["nombre"] for r in get_alumnos_usuario(u["id"])]
    nombre_real = buscar_alumno_en_lista(alumnos, nombre)
    if not nombre_real:
        return jsonify({"ok": False, "msg": f"Alumno no encontrado: {nombre}"})

    hoy = info_hoy()
    clave = f"{hoy['ano']}-{hoy['mes']:02d}"
    db_upsert_asistencia(u["id"], clave, nombre_real, str(hoy["dia"]), "J", hoy["hora"], motivo)

    regs_dia = db_get_asistencia(u["id"], clave=clave, dia=str(hoy["dia"]))
    presente = sum(1 for r in regs_dia if r["marca"] in ["A","T","J"])
    return jsonify({"ok": True, "nombre": nombre_real, "hora": hoy["hora"],
                    "presente_hoy": presente})

@app.route("/stats")
def stats():
    result = require_login()
    if isinstance(result, tuple): return result
    u = result
    if u.get("is_creator"):
        return jsonify({"presente_hoy": 0, "total_alumnos": 0, "es_hora_falta": False})

    hoy = info_hoy()
    clave = f"{hoy['ano']}-{hoy['mes']:02d}"
    regs_dia = db_get_asistencia(u["id"], clave=clave, dia=str(hoy["dia"]))
    pres = sum(1 for r in regs_dia if r["marca"] in ["A","T","J"])
    total = db_exec(
        "SELECT COUNT(*) AS c FROM alumnos WHERE usuario_id=%s", (u["id"],), fetch="one"
    )
    return jsonify({"presente_hoy": pres, "total_alumnos": total["c"] if total else 0,
                    "es_hora_falta": es_hora_falta()})

@app.route("/registros_hoy")
def registros_hoy():
    result = require_login()
    if isinstance(result, tuple): return result
    u = result

    hoy = info_hoy()
    clave = f"{hoy['ano']}-{hoy['mes']:02d}"
    regs_dia = db_get_asistencia(u["id"], clave=clave, dia=str(hoy["dia"]))

    alumnos_rows = get_alumnos_usuario(u["id"])
    genero_map = {r["nombre"]: r["genero"] for r in alumnos_rows}
    orden_map = {r["nombre"]: i for i, r in enumerate(alumnos_rows)}
    ALUMNOS = [r["nombre"] for r in alumnos_rows]

    regs = []
    for r in regs_dia:
        if r["marca"] in ["A","T","J"]:
            regs.append({"nombre": r["alumno"], "marca": r["marca"], "hora": r["hora"],
                         "genero": genero_map.get(r["alumno"], "nino"),
                         "motivo": r.get("motivo","")})
    regs.sort(key=lambda x: orden_map.get(x["nombre"], 99))
    return jsonify({"registros": regs, "total": len(ALUMNOS)})

@app.route("/faltantes_hoy")
def faltantes_hoy():
    result = require_login()
    if isinstance(result, tuple): return result
    u = result

    hoy = info_hoy()
    clave = f"{hoy['ano']}-{hoy['mes']:02d}"
    regs_dia = db_get_asistencia(u["id"], clave=clave, dia=str(hoy["dia"]))
    presentes = {r["alumno"] for r in regs_dia if r["marca"] in ["A","T","J"]}
    alumnos = [r["nombre"] for r in get_alumnos_usuario(u["id"])]
    faltantes = [a for a in alumnos if a not in presentes]
    return jsonify({"faltantes": faltantes})

@app.route("/historial")
def historial():
    result = require_login()
    if isinstance(result, tuple): return result
    u = result

    alumno = request.args.get("alumno","").strip()
    mes = int(request.args.get("mes", 0))
    ano = int(request.args.get("ano", ahora_peru().year))
    if not alumno or not mes:
        return jsonify({"dias": []})

    clave = f"{ano}-{mes:02d}"
    regs = db_get_asistencia(u["id"], clave=clave, alumno=alumno)
    regs_dict = {r["dia"]: r for r in regs}
    dias_hab = dias_habiles(ano, mes)
    resultado = []
    for dia in dias_hab:
        reg = regs_dict.get(str(dia))
        dow = calendar.weekday(ano, mes, dia)+1
        if reg:
            resultado.append({"dia": dia, "marca": reg["marca"], "hora": reg["hora"],
                              "dow": dow, "motivo": reg.get("motivo","")})
        else:
            resultado.append({"dia": dia, "marca": "?", "hora": "--:--",
                              "dow": dow, "motivo": ""})
    return jsonify({"dias": resultado})

@app.route("/resumen_meses")
def resumen_meses():
    result = require_login()
    if isinstance(result, tuple): return result
    u = result

    ano = ahora_peru().year
    mes_inicio = 1
    result_list = []
    for mes in range(mes_inicio, 13):
        clave = f"{ano}-{mes:02d}"
        dias_hab = dias_habiles(ano, mes)
        regs = db_get_asistencia(u["id"], clave=clave)
        total = sum(1 for r in regs if r["marca"] in ["A","T"])
        result_list.append({"mes": mes, "ano": ano, "nombre": MESES_ES[mes],
                            "dias_habiles": len(dias_hab), "registros": total})
    return jsonify({"meses": result_list})

@app.route("/marcar_faltas", methods=["POST"])
def api_marcar_faltas():
    result = require_login()
    if isinstance(result, tuple): return result
    u = result
    if u.get("is_creator"):
        return jsonify({"ok": True, "marcados": 0})
    return jsonify({"ok": True, "marcados": marcar_faltas_usuario(u["id"])})

@app.route("/descargar")
def descargar():
    result = require_login()
    if isinstance(result, tuple): return result
    u = result

    hoy = info_hoy()
    ano = int(request.args.get("ano", hoy["ano"]))
    mes = int(request.args.get("mes", hoy["mes"]))
    buf = generar_excel(u["id"], ano, mes, u.get("salon",""))
    return send_file(buf, as_attachment=True,
                     download_name=f"asistencia_{u.get('salon','salon')}_{ano}_{mes:02d}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/manifest.json")
def manifest():
    m = {"name":"Asistencia QR","short_name":"Asistencia QR","start_url":"/",
         "display":"standalone","background_color":"#0d1b2a","theme_color":"#1a237e",
         "icons":[{"src":"/icon.png","sizes":"192x192","type":"image/png"}]}
    return Response(json.dumps(m), mimetype="application/manifest+json")

@app.route("/icon.png")
def icon():
    from PIL import Image, ImageDraw
    img = Image.new("RGB",(192,192),"#1a237e")
    draw = ImageDraw.Draw(img)
    m=30; s=162; c=20
    for x,y in [(m,m),(s-c,m),(m,s-c),(s-c,s-c)]:
        draw.rectangle([x,y,x+c,y+c], fill="white")
    buf = io.BytesIO()
    img.save(buf,"PNG"); buf.seek(0)
    return send_file(buf, mimetype="image/png")

@app.route("/")
def index():
    return Response(HTML, mimetype="text/html")

# ══════════════════════════════════════════════════════════════════════════════
# HTML FRONTEND COMPLETO
# ══════════════════════════════════════════════════════════════════════════════
HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<meta name="mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="theme-color" content="#1a237e">
<link rel="manifest" href="/manifest.json">
<title>Asistencia QR</title>
<style>
*{margin:0;padding:0;box-sizing:border-box;-webkit-tap-highlight-color:transparent}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#0d1b2a;color:#e0e0e0;min-height:100vh}

/* ── AUTH ── */
#auth-screen{display:flex;align-items:center;justify-content:center;min-height:100vh;padding:20px}
.auth-card{background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.1);border-radius:20px;padding:32px 28px;width:100%;max-width:400px}
.auth-logo{text-align:center;margin-bottom:24px}
.auth-logo h1{font-size:1.4rem;color:#90caf9;margin-top:8px}
.auth-logo p{font-size:.8rem;color:#546e7a;margin-top:4px}
.auth-tabs{display:flex;gap:0;margin-bottom:24px;background:rgba(255,255,255,.05);border-radius:10px;padding:4px}
.auth-tab{flex:1;padding:9px;text-align:center;border:none;background:none;color:#90a4ae;font-size:.88rem;font-weight:600;cursor:pointer;border-radius:8px;font-family:inherit;transition:all .2s}
.auth-tab.active{background:#1a237e;color:#fff}
.auth-form{display:flex;flex-direction:column;gap:12px}
.auth-form input{background:rgba(255,255,255,.06);border:1px solid rgba(255,255,255,.1);border-radius:10px;padding:12px 14px;color:#fff;font-size:.9rem;font-family:inherit;outline:none}
.auth-form input:focus{border-color:#5c6bc0}
.auth-form input::placeholder{color:#546e7a}
.auth-btn{padding:13px;background:linear-gradient(135deg,#1a237e,#3949ab);color:#fff;border:none;border-radius:10px;font-size:.95rem;font-weight:700;cursor:pointer;font-family:inherit}
.auth-btn:active{opacity:.85}
.auth-err{background:rgba(183,28,28,.2);border:1px solid rgba(239,83,80,.4);color:#ef9a9a;padding:10px 12px;border-radius:8px;font-size:.83rem;display:none}
.auth-err.show{display:block}
.auth-ok{background:rgba(46,125,50,.2);border:1px solid rgba(76,175,80,.4);color:#81c784;padding:10px 12px;border-radius:8px;font-size:.83rem;display:none}
.auth-ok.show{display:block}
.codigo-hint{font-size:.75rem;color:#78909c;text-align:center}

/* ── APP ── */
#app-screen{display:none;flex-direction:column;align-items:center}
header{width:100%;background:linear-gradient(135deg,#1a237e,#283593);padding:14px 20px;position:sticky;top:0;z-index:10}
.header-top{display:flex;align-items:center;justify-content:space-between}
.header-top h1{font-size:1rem;color:#fff;font-weight:700}
.header-right{display:flex;align-items:center;gap:10px}
.user-badge{font-size:.72rem;color:#90caf9;background:rgba(255,255,255,.1);padding:4px 10px;border-radius:20px}
.btn-logout{background:none;border:none;color:#ef9a9a;font-size:.75rem;cursor:pointer;padding:4px 8px;border-radius:6px;font-family:inherit}
.btn-logout:hover{background:rgba(239,83,80,.15)}
header p{font-size:.72rem;color:#90caf9;margin-top:3px}
.wrap{width:100%;max-width:480px;padding:14px 12px;display:flex;flex-direction:column;gap:11px;padding-bottom:60px}

/* ── ADMIN PANEL ── */
.admin-panel{background:rgba(255,152,0,.06);border:1px solid rgba(255,152,0,.25);border-radius:14px;padding:16px}
.admin-title{font-size:.85rem;font-weight:700;color:#ffb74d;margin-bottom:14px;display:flex;align-items:center;gap:8px}
.admin-section{margin-bottom:16px}
.admin-section-title{font-size:.72rem;color:#78909c;text-transform:uppercase;letter-spacing:.4px;font-weight:700;margin-bottom:8px}
.btn-gen-codigo{width:100%;padding:12px;background:linear-gradient(135deg,#e65100,#f57c00);color:#fff;border:none;border-radius:10px;font-size:.9rem;font-weight:700;cursor:pointer;font-family:inherit}
.codigo-display{margin-top:10px;background:rgba(255,255,255,.05);border:1px solid rgba(255,152,0,.3);border-radius:10px;padding:14px;text-align:center;display:none}
.codigo-display.show{display:block}
.codigo-val{font-size:1.8rem;font-weight:900;letter-spacing:4px;color:#ffb74d;font-family:monospace}
.codigo-info{font-size:.72rem;color:#78909c;margin-top:6px}
.usuario-card{background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.08);border-radius:10px;padding:12px;margin-bottom:8px}
.uc-top{display:flex;align-items:center;justify-content:space-between;gap:8px}
.uc-name{font-size:.9rem;font-weight:700;color:#e0e0e0}
.uc-meta{font-size:.72rem;color:#546e7a;margin-top:3px}
.uc-badge{padding:3px 10px;border-radius:20px;font-size:.7rem;font-weight:700;flex-shrink:0}
.uc-activo{background:rgba(76,175,80,.2);color:#81c784}
.uc-inactivo{background:rgba(183,28,28,.2);color:#ef9a9a}
.uc-actions{display:flex;gap:6px;margin-top:10px}
.btn-toggle{flex:1;padding:8px;border:none;border-radius:8px;font-size:.78rem;font-weight:700;cursor:pointer;font-family:inherit}
.btn-activar{background:rgba(46,125,50,.2);color:#81c784;border:1px solid rgba(76,175,80,.3)}
.btn-desactivar{background:rgba(183,28,28,.18);color:#ef9a9a;border:1px solid rgba(239,83,80,.3)}
.btn-eliminar{padding:8px 12px;background:rgba(183,28,28,.15);color:#ef9a9a;border:1px solid rgba(239,83,80,.25);border-radius:8px;font-size:.75rem;font-weight:700;cursor:pointer;font-family:inherit}
.admin-refresh{width:100%;margin-top:6px;padding:9px;background:rgba(255,255,255,.05);color:#90a4ae;border:1px solid rgba(255,255,255,.1);border-radius:9px;font-size:.82rem;cursor:pointer;font-family:inherit}

/* ── ALUMNOS GESTOR ── */
.alumno-gestor{background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.08);border-radius:12px;padding:12px}
.ag-input-row{display:flex;gap:8px;margin-bottom:8px}
.ag-input{flex:1;background:rgba(255,255,255,.06);border:1px solid rgba(255,255,255,.1);border-radius:9px;padding:9px 11px;color:#fff;font-size:.88rem;font-family:inherit;outline:none}
.ag-select{background:rgba(255,255,255,.06);border:1px solid rgba(255,255,255,.1);border-radius:9px;padding:9px;color:#fff;font-size:.85rem;font-family:inherit;outline:none}
.ag-select option{background:#1a237e}
.ag-btn{padding:9px 14px;background:#1a237e;color:#fff;border:none;border-radius:9px;font-weight:700;cursor:pointer;font-family:inherit;font-size:.85rem}
.ag-list{max-height:200px;overflow-y:auto;display:flex;flex-direction:column;gap:4px}
.ag-item{display:flex;align-items:center;justify-content:space-between;padding:7px 10px;background:rgba(255,255,255,.04);border-radius:8px;font-size:.83rem}
.ag-item-name{flex:1;color:#e0e0e0}
.ag-item-gen{font-size:.7rem;color:#546e7a;margin-right:8px}
.ag-del{background:none;border:none;color:#ef9a9a;cursor:pointer;font-size:.8rem;padding:2px 6px;border-radius:4px}
.ag-del:hover{background:rgba(239,83,80,.15)}

/* ── RESTO ESTILOS (mismos que el app original) ── */
.cam-box{background:#000;border-radius:14px;overflow:hidden;position:relative;aspect-ratio:4/3}
#video{width:100%;height:100%;object-fit:cover;display:block}
#canvas-qr{display:none;position:absolute;top:0;left:0}
.fo{position:absolute;inset:0;display:flex;align-items:center;justify-content:center;pointer-events:none}
.qf{width:55%;aspect-ratio:1;position:relative}
.qf::before{content:'';position:absolute;inset:0;border:2px solid rgba(91,107,240,.7);border-radius:8px;box-shadow:0 0 0 9999px rgba(0,0,0,.5)}
.corner{position:absolute;width:22px;height:22px}
.tl{top:0;left:0;border-top:3px solid #fff;border-left:3px solid #fff;border-radius:4px 0 0 0}
.tr{top:0;right:0;border-top:3px solid #fff;border-right:3px solid #fff;border-radius:0 4px 0 0}
.bl{bottom:0;left:0;border-bottom:3px solid #fff;border-left:3px solid #fff;border-radius:0 0 0 4px}
.br{bottom:0;right:0;border-bottom:3px solid #fff;border-right:3px solid #fff;border-radius:0 0 4px 0}
.scan-line{position:absolute;left:0;right:0;height:2px;background:linear-gradient(90deg,transparent,#5c6bc0,transparent);animation:scan 2s linear infinite}
@keyframes scan{0%{top:0}100%{top:100%}}
#sb{text-align:center;font-size:.75rem;color:#546e7a;padding:3px 0}
.btn{width:100%;padding:12px;border:none;border-radius:10px;font-size:.95rem;font-weight:700;cursor:pointer;font-family:inherit}
.on{background:linear-gradient(135deg,#1a237e,#3949ab);color:#fff}
.off{background:rgba(183,28,28,.22);color:#ef9a9a;border:1px solid rgba(239,83,80,.3)}
.notif{border-radius:12px;padding:13px 15px;display:flex;align-items:center;gap:11px;min-height:68px}
.idle{background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.07);color:#546e7a;justify-content:center;font-size:.88rem}
.ok-a{background:rgba(46,125,50,.2);border:1px solid rgba(76,175,80,.5)}
.ok-t{background:rgba(230,81,0,.18);border:1px solid rgba(255,152,0,.45)}
.warn{background:rgba(245,127,23,.15);border:1px solid rgba(255,167,38,.35);color:#ffe082}
.err{background:rgba(183,28,28,.18);border:1px solid rgba(239,83,80,.4);color:#ef9a9a}
.nb .nm{font-weight:700;font-size:1rem;color:#fff}
.nb .sb2{font-size:.78rem;opacity:.85;margin-top:2px}
.nb .hb{display:inline-block;margin-top:5px;padding:3px 10px;border-radius:20px;font-size:.77rem;font-weight:700}
.ba{background:rgba(76,175,80,.25);color:#81c784}
.bt{background:rgba(255,152,0,.25);color:#ffb74d}
.stats{display:grid;grid-template-columns:repeat(3,1fr);gap:9px}
.stat{background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.07);border-radius:10px;padding:11px 5px;text-align:center}
.sn{font-size:1.6rem;font-weight:700;line-height:1}
.sl2{font-size:.63rem;color:#90a4ae;margin-top:3px;text-transform:uppercase;letter-spacing:.3px}
.g{color:#66bb6a}.r{color:#ef5350}.b{color:#5c6bc0}
.mbox{background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.08);border-radius:12px;padding:12px}
.mlabel{font-size:.72rem;color:#78909c;margin-bottom:8px;font-weight:600;text-transform:uppercase;letter-spacing:.4px}
.mr{display:flex;gap:8px;margin-bottom:6px}
.mi{flex:1;background:rgba(255,255,255,.06);border:1px solid rgba(255,255,255,.1);border-radius:9px;padding:9px 11px;color:#fff;font-size:.88rem;font-family:inherit;outline:none}
.mi:focus{border-color:#5c6bc0}
.mi::placeholder{color:#546e7a}
.mb{padding:9px 14px;background:#1a237e;color:#fff;border:none;border-radius:9px;font-weight:700;cursor:pointer;font-size:.88rem;font-family:inherit;white-space:nowrap}
.sugs{display:none;flex-direction:column;gap:4px;max-height:150px;overflow-y:auto}
.sugs.show{display:flex}
.sug-item{padding:8px 10px;background:rgba(255,255,255,.06);border-radius:8px;font-size:.83rem;cursor:pointer;border:1px solid rgba(255,255,255,.06);-webkit-user-select:none;user-select:none;touch-action:manipulation}
.sug-nm{font-weight:600;color:#e0e0e0}
.sug-item:active{background:rgba(91,107,240,.3)!important}
.sug-hint{font-size:.72rem;color:#546e7a;margin-top:2px}
.be{width:100%;padding:10px;border-radius:10px;font-size:.9rem;font-weight:700;cursor:pointer;font-family:inherit;background:rgba(46,125,50,.18);color:#81c784;border:1px solid rgba(76,175,80,.3)}
.ib{background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.07);border-radius:10px;padding:10px;font-size:.76rem;color:#78909c;text-align:center;line-height:1.7}
.ib strong{color:#cfd8dc}
.seccion{background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.07);border-radius:12px;overflow:hidden}
.sec-btn{width:100%;padding:12px 16px;background:none;border:none;color:#e0e0e0;font-size:.88rem;font-weight:700;cursor:pointer;font-family:inherit;display:flex;align-items:center;justify-content:space-between}
.sec-btn:active{background:rgba(255,255,255,.04)}
.sec-btn .arrow{transition:transform .3s;font-size:.75rem;color:#546e7a}
.sec-btn.open .arrow{transform:rotate(180deg)}
.sec-content{display:none;padding:0 12px 12px}
.sec-content.show{display:block}
.li{display:flex;align-items:center;gap:9px;padding:7px 0;border-bottom:1px solid rgba(255,255,255,.05)}
.li:last-child{border:none}
.av{width:32px;height:32px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-weight:700;font-size:.8rem;color:#fff;flex-shrink:0}
.ava{background:linear-gradient(135deg,#2e7d32,#43a047)}
.avt{background:linear-gradient(135deg,#e65100,#f57c00)}
.avf{background:linear-gradient(135deg,#b71c1c,#c62828)}
.avj{background:linear-gradient(135deg,#1565c0,#1976d2)}
.ln{font-size:.83rem;font-weight:600}
.ls{font-size:.67rem;color:#546e7a;margin-top:1px}
.rb{margin-left:auto;border-radius:20px;padding:2px 9px;font-size:.67rem;font-weight:700;white-space:nowrap}
.rba{background:rgba(76,175,80,.2);color:#81c784}
.rbt{background:rgba(255,152,0,.2);color:#ffb74d}
.rbf{background:rgba(183,28,28,.2);color:#ef9a9a}
.rbj{background:rgba(91,107,240,.2);color:#9fa8da}
.just-box{background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.08);border-radius:10px;padding:12px}
.just-label{font-size:.72rem;color:#78909c;margin-bottom:8px;font-weight:600;text-transform:uppercase;letter-spacing:.4px}
.just-textarea{width:100%;background:rgba(255,255,255,.06);border:1px solid rgba(255,255,255,.1);border-radius:9px;padding:9px 11px;color:#fff;font-size:.85rem;font-family:inherit;outline:none;resize:vertical;min-height:70px;margin-top:6px}
.just-textarea::placeholder{color:#546e7a}
.btn-just{width:100%;margin-top:8px;padding:10px;background:linear-gradient(135deg,#1565c0,#1976d2);color:#fff;border:none;border-radius:9px;font-weight:700;cursor:pointer;font-family:inherit;font-size:.88rem}
.hist-sel{width:100%;background:rgba(255,255,255,.06);border:1px solid rgba(255,255,255,.1);border-radius:9px;padding:9px 11px;color:#fff;font-size:.88rem;font-family:inherit;outline:none;margin-bottom:8px}
.hist-sel option{background:#1a237e;color:#fff}
.hist-row{display:flex;align-items:center;gap:8px;padding:6px 8px;border-radius:6px;margin-bottom:3px}
.hist-dia{width:28px;text-align:center;font-weight:700;color:#90a4ae;flex-shrink:0;font-size:.82rem}
.hist-letra{width:24px;text-align:center;flex-shrink:0;font-weight:700}
.hist-hora{margin-left:auto;font-size:.72rem;color:#546e7a}
.hist-resumen{display:flex;gap:10px;margin-bottom:10px;flex-wrap:wrap}
.hist-badge{padding:4px 12px;border-radius:20px;font-size:.75rem;font-weight:700}
.empty-msg{color:#546e7a;font-size:.8rem;text-align:center;padding:10px 0}
.meses-grid{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:6px}
.mes-card{background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.08);border-radius:10px;padding:10px 12px;display:flex;align-items:center;justify-content:space-between;gap:8px}
.mes-nombre{font-size:.85rem;font-weight:700;color:#e0e0e0}
.mes-regs{font-size:.68rem;margin-top:3px}
.mes-tiene{color:#81c784}
.mes-vacio{color:#546e7a}
.btn-dl-mes{padding:6px 10px;background:rgba(46,125,50,.2);color:#81c784;border:1px solid rgba(76,175,80,.3);border-radius:8px;font-size:.75rem;font-weight:700;cursor:pointer;font-family:inherit;white-space:nowrap}
.genero-grid{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-top:4px}
.genero-col{background:rgba(255,255,255,.03);border-radius:10px;padding:10px}
.genero-titulo{font-size:.75rem;font-weight:700;text-align:center;margin-bottom:8px;padding-bottom:6px;border-bottom:1px solid rgba(255,255,255,.08)}
.titulo-nino{color:#64b5f6}
.titulo-nina{color:#f48fb1}
.alumno-item{display:flex;align-items:center;gap:6px;padding:4px 0;border-bottom:1px solid rgba(255,255,255,.04);font-size:.76rem}
.alumno-item:last-child{border:none}
.grupos-container{display:flex;flex-direction:column;gap:9px;margin-top:8px}
.grupo-card{border-radius:10px;padding:10px}
.grupo-titulo{font-size:.77rem;font-weight:700;margin-bottom:7px;text-align:center}
.grupo-miembros{display:flex;flex-wrap:wrap;gap:5px}
.miembro{display:flex;align-items:center;gap:4px;background:rgba(255,255,255,.07);border-radius:20px;padding:3px 9px;font-size:.73rem}
.input-grupos{display:flex;gap:8px;margin-top:4px}
.input-num{width:70px;background:rgba(255,255,255,.06);border:1px solid rgba(255,255,255,.1);border-radius:9px;padding:8px;color:#fff;font-size:.9rem;font-family:inherit;outline:none;text-align:center}
.btn-formar{flex:1;padding:9px;background:linear-gradient(135deg,#1a237e,#3949ab);color:#fff;border:none;border-radius:9px;font-weight:700;cursor:pointer;font-family:inherit;font-size:.85rem}
.btn-regen{width:100%;padding:8px;background:rgba(91,107,240,.18);color:#7986cb;border:1px solid rgba(91,107,240,.3);border-radius:9px;font-weight:700;cursor:pointer;font-family:inherit;font-size:.83rem;margin-top:7px;display:none}
</style>
</head>
<body>

<!-- ═══ PANTALLA AUTH ═══ -->
<div id="auth-screen">
<div class="auth-card">
  <div class="auth-logo">
    <div style="font-size:2.5rem">🏫</div>
    <h1>Asistencia QR</h1>
    <p>Sistema de control de asistencia</p>
  </div>
  <div class="auth-tabs">
    <button class="auth-tab active" onclick="authTab('login')">Iniciar Sesión</button>
    <button class="auth-tab" onclick="authTab('registro')">Registrarse</button>
  </div>

  <!-- Login -->
  <div id="form-login" class="auth-form">
    <input id="l-user" placeholder="Usuario" autocomplete="username" autocorrect="off" autocapitalize="off">
    <input id="l-pass" type="password" placeholder="Contraseña" autocomplete="current-password">
    <div id="l-err" class="auth-err"></div>
    <button class="auth-btn" onclick="doLogin()">Entrar →</button>
  </div>

  <!-- Registro -->
  <div id="form-registro" class="auth-form" style="display:none">
    <input id="r-codigo" placeholder="Código de invitación (10 dígitos)" autocorrect="off" autocapitalize="characters" spellcheck="false" style="font-family:monospace;letter-spacing:2px;text-transform:uppercase">
    <p class="codigo-hint">⚠️ El código debe dártelo el administrador. Válido 10 min, un solo uso.</p>
    <input id="r-user" placeholder="Nombre de usuario (sin espacios)" autocomplete="username" autocorrect="off" autocapitalize="off">
    <input id="r-pass" type="password" placeholder="Contraseña (mín. 6 caracteres)">
    <input id="r-nombre" placeholder="Tu nombre completo" autocorrect="off">
    <input id="r-inst" placeholder="Institución educativa" autocorrect="off">
    <input id="r-salon" placeholder="Salón (ej: 3°F, 2°A)" autocorrect="off">
    <div id="r-err" class="auth-err"></div>
    <div id="r-ok" class="auth-ok"></div>
    <button class="auth-btn" onclick="doRegistro()">Crear cuenta →</button>
  </div>
</div>
</div>

<!-- ═══ PANTALLA APP ═══ -->
<div id="app-screen">
<header>
  <div class="header-top">
    <h1>🏫 <span id="h-salon">Asistencia QR</span></h1>
    <div class="header-right">
      <span class="user-badge" id="h-user">...</span>
      <button class="btn-logout" onclick="doLogout()">Salir</button>
    </div>
  </div>
  <p id="clk"></p>
</header>

<div class="wrap" id="main-wrap">

  <!-- PANEL ADMIN -->
  <div id="admin-panel" class="admin-panel" style="display:none">
    <div class="admin-title">⚙️ Panel Administrador — Creador</div>

    <div class="admin-section">
      <div class="admin-section-title">🔑 Código de Invitación</div>
      <button class="btn-gen-codigo" onclick="generarCodigo()">Generar nuevo código (10 min)</button>
      <div id="codigo-display" class="codigo-display">
        <div class="codigo-val" id="codigo-val">—</div>
        <div class="codigo-info">⏱️ Expira: <span id="codigo-expira">—</span> | Un solo uso</div>
        <div style="margin-top:10px">
          <button onclick="copiarCodigo()" style="background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.15);color:#e0e0e0;padding:7px 16px;border-radius:8px;cursor:pointer;font-size:.82rem;font-family:inherit">📋 Copiar código</button>
        </div>
      </div>
    </div>

    <div class="admin-section">
      <div class="admin-section-title">👤 Usuarios Registrados</div>
      <div id="admin-usuarios-list"><div class="empty-msg">Cargando...</div></div>
      <button class="admin-refresh" onclick="cargarUsuariosAdmin()">🔄 Actualizar lista</button>
    </div>
  </div>

  <!-- VISTA NORMAL (profesores) -->
  <div id="vista-normal">
    <div class="ib">
      🟢 <strong>A</strong> = Antes 7:56 &nbsp; 🟠 <strong>T</strong> = Tardanza &nbsp; 🔴 <strong>F</strong> = Falta (4 PM)
    </div>

    <div class="cam-box">
      <video id="video" playsinline autoplay muted webkit-playsinline></video>
      <canvas id="canvas-qr"></canvas>
      <button id="btnFlip" onclick="voltearCamara()" style="display:none;position:absolute;bottom:10px;right:10px;background:rgba(0,0,0,.55);border:none;border-radius:50%;width:42px;height:42px;font-size:1.25rem;cursor:pointer;color:#fff;z-index:5">🔄</button>
      <div class="fo"><div class="qf">
        <div class="scan-line"></div>
        <div class="corner tl"></div><div class="corner tr"></div>
        <div class="corner bl"></div><div class="corner br"></div>
      </div></div>
    </div>
    <div id="sb">Cámara detenida</div>
    <button id="btnCam" class="btn on" onclick="toggleCam()">📷 Iniciar Cámara</button>
    <div id="notif" class="notif idle">Esperando escaneo...</div>

    <div class="stats">
      <div class="stat"><div class="sn g" id="sp">0</div><div class="sl2">Presentes</div></div>
      <div class="stat"><div class="sn r" id="sa">-</div><div class="sl2">Ausentes</div></div>
      <div class="stat"><div class="sn b" id="sc">-</div><div class="sl2">% Hoy</div></div>
    </div>

    <!-- Registro manual -->
    <div class="mbox">
      <div class="mlabel">✏️ Registro manual</div>
      <div class="mr">
        <input id="mi" class="mi" placeholder="Escribe nombre o apellido..." oninput="buscarAlumnoManual(this.value)" autocomplete="off" autocorrect="off" autocapitalize="off" spellcheck="false">
        <button class="mb" onclick="regManualSel()">✓</button>
      </div>
      <div id="sugs" class="sugs"></div>
    </div>

    <button class="be" onclick="dlExcel()">📊 Descargar Excel del mes actual</button>

    <!-- Justificación -->
    <div class="seccion">
      <button class="sec-btn" onclick="toggleSec('sec-just',this)">
        📝 Registrar justificación <span class="arrow">▼</span>
      </button>
      <div id="sec-just" class="sec-content">
        <div class="just-box">
          <div class="just-label">Buscar alumno</div>
          <div class="mr" style="margin-bottom:6px">
            <input id="mi-just" class="mi" placeholder="Nombre o apellido..." oninput="buscarJust(this.value)" autocomplete="off" autocorrect="off" autocapitalize="off" spellcheck="false">
          </div>
          <div id="sugs-just" class="sugs"></div>
          <textarea id="texto-just" class="just-textarea" placeholder="Motivo (ej: cita médica, viaje familiar...)"></textarea>
          <button class="btn-just" onclick="guardarJustificacion()">Guardar justificación</button>
          <div id="just-result"></div>
        </div>
      </div>
    </div>

    <!-- Registros hoy -->
    <div class="seccion">
      <button class="sec-btn" onclick="toggleSec('sec-reg',this);cargarRegistros()">
        🕐 Registros de hoy <span class="arrow">▼</span>
      </button>
      <div id="sec-reg" class="sec-content">
        <div id="lista-reg"><div class="empty-msg">Presiona para ver</div></div>
      </div>
    </div>

    <!-- Faltantes -->
    <div class="seccion">
      <button class="sec-btn" onclick="toggleSec('sec-falt',this);cargarFaltantes()">
        🚫 Alumnos faltantes hoy <span class="arrow">▼</span>
      </button>
      <div id="sec-falt" class="sec-content">
        <div id="lista-falt"><div class="empty-msg">Presiona para ver</div></div>
      </div>
    </div>

    <!-- Género -->
    <div class="seccion">
      <button class="sec-btn" onclick="toggleSec('sec-gen',this);cargarGenero()">
        👦👧 Alumnos por género <span class="arrow">▼</span>
      </button>
      <div id="sec-gen" class="sec-content">
        <div id="gen-content"><div class="empty-msg">Presiona para ver</div></div>
      </div>
    </div>

    <!-- Grupos -->
    <div class="seccion">
      <button class="sec-btn" onclick="toggleSec('sec-grp',this)">
        🎲 Grupos aleatorios <span class="arrow">▼</span>
      </button>
      <div id="sec-grp" class="sec-content">
        <div class="input-grupos">
          <input id="num-grp" class="input-num" type="number" min="2" max="10" value="4" inputmode="numeric">
          <button class="btn-formar" onclick="formarGrupos()">👥 Formar grupos</button>
        </div>
        <div id="grp-result"></div>
        <button class="btn-regen" id="btn-regen" onclick="formarGrupos()">🔄 Regenerar</button>
      </div>
    </div>

    <!-- Historial -->
    <div class="seccion">
      <button class="sec-btn" onclick="toggleSec('sec-hist',this);initHistorial()">
        📋 Historial por estudiante <span class="arrow">▼</span>
      </button>
      <div id="sec-hist" class="sec-content">
        <select class="hist-sel" id="hist-alumno" onchange="histCambioAlumno()">
          <option value="">-- Selecciona un alumno --</option>
        </select>
        <select class="hist-sel" id="hist-mes" onchange="cargarHistorial()" style="display:none">
          <option value="">-- Selecciona un mes --</option>
        </select>
        <div id="hist-result"><div class="empty-msg">Selecciona un alumno y un mes</div></div>
      </div>
    </div>

    <!-- Excel por mes -->
    <div class="seccion">
      <button class="sec-btn" onclick="toggleSec('sec-meses',this);cargarMeses()">
        📅 Descargar Excel por mes <span class="arrow">▼</span>
      </button>
      <div id="sec-meses" class="sec-content">
        <div id="meses-grid" class="meses-grid"><div class="empty-msg">Cargando...</div></div>
      </div>
    </div>

    <!-- Mis alumnos -->
    <div class="seccion">
      <button class="sec-btn" onclick="toggleSec('sec-alumnos',this);cargarAlumnosGestor()">
        🎓 Gestionar mis alumnos <span class="arrow">▼</span>
      </button>
      <div id="sec-alumnos" class="sec-content">
        <div class="alumno-gestor">
          <div class="mlabel">Agregar alumno</div>
          <div class="ag-input-row">
            <input id="ag-nombre" class="ag-input" placeholder="Nombre completo" autocorrect="off" autocapitalize="words">
            <select id="ag-genero" class="ag-select">
              <option value="nino">👦</option>
              <option value="nina">👧</option>
            </select>
            <button class="ag-btn" onclick="agregarAlumno()">+</button>
          </div>
          <div id="ag-msg" style="font-size:.78rem;color:#ef9a9a;margin-bottom:6px;display:none"></div>
          <div class="mlabel" style="margin-top:8px">Lista de alumnos</div>
          <div id="ag-list" class="ag-list"><div class="empty-msg">Sin alumnos aún</div></div>
        </div>
      </div>
    </div>
  </div><!-- /vista-normal -->

</div><!-- /wrap -->
</div><!-- /app-screen -->

<script>
// ═══════════════════════════════════════════════════════
// ESTADO GLOBAL
// ═══════════════════════════════════════════════════════
let currentUser = null;
let ALUMNOS_JS = [];
let alumnoSel = null, alumnoJustSel = null;
let camaraActual = 'environment', streamActual = null;
let camOn = false, cooldown = false, iId = null;
const esIOS = /iPad|iPhone|iPod/.test(navigator.userAgent)||(navigator.platform==='MacIntel'&&navigator.maxTouchPoints>1);
const tieneBarcode = 'BarcodeDetector' in window;
const MESES_NOMBRES=['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE'];

// ═══════════════════════════════════════════════════════
// AUTH
// ═══════════════════════════════════════════════════════
function authTab(tab) {
  document.querySelectorAll('.auth-tab').forEach(t=>t.classList.remove('active'));
  event.target.classList.add('active');
  document.getElementById('form-login').style.display = tab==='login'?'flex':'none';
  document.getElementById('form-registro').style.display = tab==='registro'?'flex':'none';
}

async function doLogin() {
  const username = document.getElementById('l-user').value.trim();
  const password = document.getElementById('l-pass').value.trim();
  const errEl = document.getElementById('l-err');
  errEl.className='auth-err';
  if(!username||!password){errEl.textContent='Completa todos los campos';errEl.className='auth-err show';return;}
  try{
    const r = await fetch('/auth/login',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({username,password}),credentials:'include'});
    const d = await r.json();
    if(d.ok){
      currentUser = d;
      iniciarApp(d);
    } else {
      errEl.textContent = d.msg; errEl.className='auth-err show';
    }
  }catch(e){errEl.textContent='Sin conexión';errEl.className='auth-err show';}
}

async function doRegistro() {
  const codigo = document.getElementById('r-codigo').value.trim().toUpperCase();
  const username = document.getElementById('r-user').value.trim();
  const password = document.getElementById('r-pass').value.trim();
  const nombre_completo = document.getElementById('r-nombre').value.trim();
  const institucion = document.getElementById('r-inst').value.trim();
  const salon = document.getElementById('r-salon').value.trim();
  const errEl = document.getElementById('r-err');
  const okEl = document.getElementById('r-ok');
  errEl.className='auth-err'; okEl.className='auth-ok';
  try{
    const r = await fetch('/auth/registro',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({codigo,username,password,nombre_completo,institucion,salon}),credentials:'include'});
    const d = await r.json();
    if(d.ok){
      currentUser = d;
      okEl.textContent='¡Cuenta creada! Entrando...'; okEl.className='auth-ok show';
      setTimeout(()=>iniciarApp(d), 800);
    } else {
      errEl.textContent = d.msg; errEl.className='auth-err show';
    }
  }catch(e){errEl.textContent='Sin conexión';errEl.className='auth-err show';}
}

async function doLogout(){
  await fetch('/auth/logout',{method:'POST',credentials:'include'});
  session.clear?.();
  location.reload();
}

async function checkAuth(){
  try{
    const r = await fetch('/auth/me',{credentials:'include'});
    const d = await r.json();
    if(d.ok){ currentUser=d; iniciarApp(d); }
    else { document.getElementById('auth-screen').style.display='flex'; }
  }catch(e){ document.getElementById('auth-screen').style.display='flex'; }
}

function iniciarApp(user){
  document.getElementById('auth-screen').style.display='none';
  const appEl = document.getElementById('app-screen');
  appEl.style.display='flex';

  document.getElementById('h-user').textContent = user.username || 'Usuario';
  document.getElementById('h-salon').textContent = user.is_creator ? '⚙️ Admin' : (user.salon||'Asistencia QR');

  if(user.is_creator){
    document.getElementById('admin-panel').style.display='block';
    document.getElementById('vista-normal').style.display='none';
    cargarUsuariosAdmin();
  } else {
    document.getElementById('admin-panel').style.display='none';
    document.getElementById('vista-normal').style.display='block';
    cargarAlumnosJS();
    fetch('/stats',{credentials:'include'}).then(r=>r.json()).then(d=>{
      stats(d.presente_hoy, d.total_alumnos);
      if(d.es_hora_falta) fetch('/marcar_faltas',{method:'POST',credentials:'include'}).catch(()=>{});
    }).catch(()=>{});
  }
  setInterval(tick, 1000); tick();
}

async function cargarAlumnosJS(){
  try{
    const r = await fetch('/alumnos/lista',{credentials:'include'});
    const d = await r.json();
    if(d.ok){ ALUMNOS_JS = d.alumnos.map(a=>a.nombre); }
  }catch(e){}
}

// ═══════════════════════════════════════════════════════
// ADMIN
// ═══════════════════════════════════════════════════════
async function generarCodigo(){
  try{
    const r = await fetch('/admin/generar_codigo',{method:'POST',credentials:'include'});
    const d = await r.json();
    if(d.ok){
      document.getElementById('codigo-val').textContent = d.codigo;
      document.getElementById('codigo-expira').textContent = d.expira_en;
      document.getElementById('codigo-display').className='codigo-display show';
    }
  }catch(e){alert('Error generando código');}
}

function copiarCodigo(){
  const c = document.getElementById('codigo-val').textContent;
  navigator.clipboard.writeText(c).then(()=>alert('Código copiado: '+c));
}

async function cargarUsuariosAdmin(){
  try{
    const r = await fetch('/admin/usuarios',{credentials:'include'});
    const d = await r.json();
    if(!d.ok) return;
    const el = document.getElementById('admin-usuarios-list');
    if(!d.usuarios.length){
      el.innerHTML='<div class="empty-msg">No hay usuarios registrados aún</div>';
      return;
    }
    el.innerHTML = d.usuarios.map(u=>`
      <div class="usuario-card">
        <div class="uc-top">
          <div>
            <div class="uc-name">👤 ${u.username}</div>
            <div class="uc-meta">${u.nombre_completo||''} · ${u.salon||'Sin salón'} · ${u.institucion||''}</div>
            <div class="uc-meta" style="margin-top:2px">📚 ${u.num_alumnos} alumnos · Registrado: ${u.creado_en}</div>
          </div>
          <span class="uc-badge ${u.activo?'uc-activo':'uc-inactivo'}">${u.activo?'Activo':'Inactivo'}</span>
        </div>
        <div class="uc-actions">
          ${u.activo
            ? `<button class="btn-toggle btn-desactivar" onclick="toggleUsuario(${u.id},false)">🔒 Desactivar</button>`
            : `<button class="btn-toggle btn-activar" onclick="toggleUsuario(${u.id},true)">✅ Activar</button>`
          }
          <button class="btn-eliminar" onclick="eliminarUsuario(${u.id},'${u.username}')">🗑</button>
        </div>
      </div>
    `).join('');
  }catch(e){}
}

async function toggleUsuario(id, activo){
  try{
    const r = await fetch('/admin/toggle_usuario',{method:'POST',credentials:'include',
      headers:{'Content-Type':'application/json'},body:JSON.stringify({id,activo})});
    const d = await r.json();
    if(d.ok) cargarUsuariosAdmin();
    else alert(d.msg);
  }catch(e){}
}

async function eliminarUsuario(id, username){
  if(!confirm(`¿Eliminar al usuario "${username}" y todos sus datos? Esta acción no se puede deshacer.`)) return;
  try{
    const r = await fetch('/admin/eliminar_usuario',{method:'POST',credentials:'include',
      headers:{'Content-Type':'application/json'},body:JSON.stringify({id})});
    const d = await r.json();
    if(d.ok) cargarUsuariosAdmin();
    else alert(d.msg);
  }catch(e){}
}

// ═══════════════════════════════════════════════════════
// ALUMNOS (gestor)
// ═══════════════════════════════════════════════════════
async function cargarAlumnosGestor(){
  try{
    const r = await fetch('/alumnos/lista',{credentials:'include'});
    const d = await r.json();
    if(!d.ok) return;
    const el = document.getElementById('ag-list');
    if(!d.alumnos.length){el.innerHTML='<div class="empty-msg">Sin alumnos aún</div>';return;}
    el.innerHTML = d.alumnos.map(a=>`
      <div class="ag-item">
        <span class="ag-item-name">${a.nombre}</span>
        <span class="ag-item-gen">${a.genero==='nino'?'👦':'👧'}</span>
        <button class="ag-del" onclick="eliminarAlumno('${a.nombre.replace(/'/g,"\\'")}')">✕</button>
      </div>
    `).join('');
  }catch(e){}
}

async function agregarAlumno(){
  const nombre = document.getElementById('ag-nombre').value.trim();
  const genero = document.getElementById('ag-genero').value;
  const msgEl = document.getElementById('ag-msg');
  msgEl.style.display='none';
  if(!nombre){msgEl.textContent='Escribe el nombre';msgEl.style.display='block';return;}
  try{
    const r = await fetch('/alumnos/agregar',{method:'POST',credentials:'include',
      headers:{'Content-Type':'application/json'},body:JSON.stringify({nombre,genero})});
    const d = await r.json();
    if(d.ok){
      document.getElementById('ag-nombre').value='';
      cargarAlumnosGestor();
      cargarAlumnosJS();
    } else {
      msgEl.textContent=d.msg; msgEl.style.display='block';
    }
  }catch(e){}
}

async function eliminarAlumno(nombre){
  if(!confirm(`¿Eliminar a "${nombre}"?`)) return;
  try{
    await fetch('/alumnos/eliminar',{method:'POST',credentials:'include',
      headers:{'Content-Type':'application/json'},body:JSON.stringify({nombre})});
    cargarAlumnosGestor();
    cargarAlumnosJS();
  }catch(e){}
}

// ═══════════════════════════════════════════════════════
// RELOJ
// ═══════════════════════════════════════════════════════
function tick(){
  const n=new Date();
  const el=document.getElementById('clk');
  if(el) el.textContent=n.toLocaleDateString('es-PE',{weekday:'short',day:'numeric',month:'short'})+' '+n.toLocaleTimeString('es-PE',{hour:'2-digit',minute:'2-digit',second:'2-digit'});
  if(n.getHours()===16&&n.getMinutes()===0&&n.getSeconds()===0)
    fetch('/marcar_faltas',{method:'POST',credentials:'include'}).catch(()=>{});
}

// ═══════════════════════════════════════════════════════
// BÚSQUEDA
// ═══════════════════════════════════════════════════════
function norm(s){return s.toLowerCase().normalize('NFD').replace(/[\u0300-\u036f]/g,'');}

function mostrarSugs(q, listaId, onSelectFn){
  const sug=document.getElementById(listaId);
  if(!q||q.length<1){sug.className='sugs';sug.innerHTML='';return;}
  const qn=norm(q);
  const matches=ALUMNOS_JS.filter(a=>norm(a).includes(qn)).slice(0,8);
  if(!matches.length){sug.className='sugs';sug.innerHTML='';return;}
  sug.className='sugs show';
  sug.innerHTML=matches.map(a=>{
    const p=a.split(',');const ap=p[0]||'';const nm=p[1]?p[1].trim():'';
    const label=nm+' '+ap;
    return `<div class="sug-item" onmousedown="event.preventDefault()" onclick="(${onSelectFn.toString()})(${JSON.stringify(a)})"><div class="sug-nm">${label}</div><div class="sug-hint">${a}</div></div>`;
  }).join('');
}

function buscarAlumnoManual(q){
  alumnoSel=null;
  mostrarSugs(q,'sugs',function(nombre){
    alumnoSel=nombre;
    document.getElementById('mi').value=nombre;
    document.getElementById('sugs').className='sugs';
    document.getElementById('sugs').innerHTML='';
  });
}

function regManualSel(){
  const nombre=alumnoSel||document.getElementById('mi').value.trim();
  if(!nombre) return;
  enviar(nombre);
  document.getElementById('mi').value='';
  alumnoSel=null;
  document.getElementById('sugs').className='sugs';
  document.getElementById('sugs').innerHTML='';
}

function buscarJust(q){
  alumnoJustSel=null;
  mostrarSugs(q,'sugs-just',function(nombre){
    alumnoJustSel=nombre;
    document.getElementById('mi-just').value=nombre;
    document.getElementById('sugs-just').className='sugs';
    document.getElementById('sugs-just').innerHTML='';
    document.getElementById('texto-just').focus();
  });
}

async function guardarJustificacion(){
  const nombre=alumnoJustSel||document.getElementById('mi-just').value.trim();
  const motivo=document.getElementById('texto-just').value.trim();
  if(!nombre){alert('Selecciona un alumno');return;}
  if(!motivo){alert('Escribe el motivo');return;}
  try{
    const r=await fetch('/justificar',{method:'POST',credentials:'include',
      headers:{'Content-Type':'application/json'},body:JSON.stringify({nombre,motivo})});
    const d=await r.json();
    if(d.ok){
      document.getElementById('mi-just').value='';
      document.getElementById('texto-just').value='';
      alumnoJustSel=null;
      document.getElementById('just-result').innerHTML='<div style="color:#81c784;font-size:.85rem;padding:8px 0">✅ Justificación guardada para '+d.nombre+'</div>';
      setTimeout(()=>{document.getElementById('just-result').innerHTML='';},4000);
      stats(d.presente_hoy, ALUMNOS_JS.length);
    }else{
      document.getElementById('just-result').innerHTML='<div style="color:#ef9a9a;font-size:.83rem;padding:6px 0">❌ '+d.msg+'</div>';
    }
  }catch(e){}
}

// ═══════════════════════════════════════════════════════
// CÁMARA
// ═══════════════════════════════════════════════════════
async function cargarJsQR(){
  if(window.jsQR) return;
  await new Promise(res=>{
    const s=document.createElement('script');
    s.src='https://cdnjs.cloudflare.com/ajax/libs/jsQR/1.4.0/jsQR.min.js';
    s.onload=()=>res(); s.onerror=res;
    document.head.appendChild(s);
  });
}

async function iniciarStream(facing){
  if(streamActual){streamActual.getTracks().forEach(t=>t.stop());streamActual=null;}
  const constraints=esIOS?{video:{facingMode:facing}}:{video:{facingMode:{ideal:facing},width:{ideal:1280},height:{ideal:720}}};
  const stream=await navigator.mediaDevices.getUserMedia(constraints);
  streamActual=stream;
  const v=document.getElementById('video');
  v.srcObject=stream; v.muted=true; v.playsInline=true;
  await new Promise(res=>{
    let done=false; const finish=()=>{if(!done){done=true;res();}};
    v.onloadeddata=finish; v.oncanplay=finish;
    v.onloadedmetadata=()=>{v.play().then(finish).catch(finish);};
    setTimeout(finish,5000);
  });
  for(let i=0;i<5;i++){try{await v.play();if(!v.paused)break;}catch(e){} await new Promise(r=>setTimeout(r,300));}
}

function iniciarEscaner(){
  if(iId){cancelAnimationFrame(iId);clearInterval(iId);iId=null;}
  const v=document.getElementById('video');
  if(tieneBarcode&&!esIOS){
    const det=new BarcodeDetector({formats:['qr_code']});
    iId=setInterval(async()=>{
      if(!camOn) return;
      try{const codes=await det.detect(v);if(codes.length>0&&!cooldown){cooldown=true;procQR(codes[0].rawValue);setTimeout(()=>cooldown=false,2500);}}catch(e){}
    },250);
    return;
  }
  const canvas=document.getElementById('canvas-qr');
  const ctx=canvas.getContext('2d',{willReadFrequently:true});
  let intentos=0;
  iId=setInterval(()=>{
    if(!camOn) return; intentos++;
    if(v.paused||v.ended){v.play().catch(()=>{});return;}
    if(v.readyState<2||v.videoWidth===0||v.videoHeight===0) return;
    try{
      const W=Math.min(v.videoWidth,480),H=Math.min(v.videoHeight,360);
      canvas.width=W; canvas.height=H;
      ctx.drawImage(v,0,0,W,H);
      if(!window.jsQR) return;
      const imgData=ctx.getImageData(0,0,W,H);
      let qr=window.jsQR(imgData.data,imgData.width,imgData.height,{inversionAttempts:'dontInvert'});
      if(!qr) qr=window.jsQR(imgData.data,imgData.width,imgData.height,{inversionAttempts:'onlyInvert'});
      if(qr&&qr.data&&!cooldown){cooldown=true;procQR(qr.data);setTimeout(()=>cooldown=false,2500);}
    }catch(e){}
  },100);
}

async function toggleCam(){
  if(camOn){stopCam();return;}
  await cargarJsQR();
  try{
    document.getElementById('sb').textContent='Activando cámara...';
    await iniciarStream(camaraActual);
    camOn=true;
    document.getElementById('btnCam').textContent='⏹️ Detener Cámara';
    document.getElementById('btnCam').className='btn off';
    document.getElementById('sb').textContent=esIOS?'📷 Acerca el QR a la cámara':'🟢 Apunta el QR del alumno';
    document.getElementById('btnFlip').style.display='block';
    iniciarEscaner();
  }catch(e){
    setN('err','Sin acceso a cámara',esIOS?'Ve a Ajustes → Safari → Cámara → Permitir':'Permite el permiso de cámara');
    document.getElementById('sb').textContent='Error de cámara';
  }
}

async function voltearCamara(){
  if(!camOn) return;
  camaraActual=camaraActual==='environment'?'user':'environment';
  if(iId){cancelAnimationFrame(iId);clearInterval(iId);iId=null;}
  try{await iniciarStream(camaraActual);iniciarEscaner();}
  catch(e){camaraActual=camaraActual==='environment'?'user':'environment';}
}

function stopCam(){
  camOn=false;
  if(iId){cancelAnimationFrame(iId);clearInterval(iId);iId=null;}
  if(streamActual){streamActual.getTracks().forEach(t=>t.stop());streamActual=null;}
  document.getElementById('video').srcObject=null;
  document.getElementById('btnCam').textContent='📷 Iniciar Cámara';
  document.getElementById('btnCam').className='btn on';
  document.getElementById('sb').textContent='Cámara detenida';
  document.getElementById('btnFlip').style.display='none';
  camaraActual='environment';
}

function procQR(d){
  document.getElementById('sb').textContent='QR detectado!';
  const p=d.split('|');
  if(p.length!==3||p[0]!=='ASIST'){setN('err','QR inválido','No pertenece al sistema');return;}
  enviar(p[2]);
}

async function enviar(nombre){
  try{
    const r=await fetch('/registrar',{method:'POST',credentials:'include',
      headers:{'Content-Type':'application/json'},body:JSON.stringify({nombre})});
    const d=await r.json();
    if(d.ok){
      const a=d.tipo==='temprano';
      setN(a?'ok-a':'ok-t',d.nombre,a?'✅ Asistió a tiempo':'⚠️ Tardanza registrada',d.hora,a);
      stats(d.presente_hoy, ALUMNOS_JS.length);
      if(document.getElementById('sec-falt').classList.contains('show')) cargarFaltantes();
    }else if(d.msg==='ya_registrado'){
      setN('warn',d.nombre,'Ya fue registrado hoy');
    }else{setN('err','Error',d.msg);}
  }catch(e){setN('err','Sin conexión','Verifica tu internet');}
}

let nt=null;
function setN(cls,nm,sb,hora,a){
  const el=document.getElementById('notif');
  el.className='notif '+cls;
  const h=hora?`<div class="hb ${a?'ba':'bt'}">🕐 ${hora} — ${a?'A tiempo':'Tardanza'}</div>`:'';
  el.innerHTML=`<div class="nb"><div class="nm">${nm}</div><div class="sb2">${sb}</div>${h}</div>`;
  if(nt) clearTimeout(nt);
  nt=setTimeout(()=>{el.className='notif idle';el.textContent='Esperando escaneo...';},6000);
}

function stats(p, total){
  const tot=total!==undefined?total:ALUMNOS_JS.length;
  const pr=p!==undefined?p:0;
  document.getElementById('sp').textContent=pr;
  document.getElementById('sa').textContent=tot-pr;
  document.getElementById('sc').textContent=tot>0?Math.round(pr/tot*100)+'%':'-';
}

// ═══════════════════════════════════════════════════════
// SECCIONES
// ═══════════════════════════════════════════════════════
function toggleSec(id,btn){
  const el=document.getElementById(id);
  const open=el.classList.toggle('show');
  btn.classList.toggle('open',open);
}

async function cargarRegistros(){
  try{
    const r=await fetch('/registros_hoy',{credentials:'include'});const d=await r.json();
    const el=document.getElementById('lista-reg');
    if(!d.registros.length){el.innerHTML='<div class="empty-msg">Ningún registro aún</div>';return;}
    el.innerHTML=d.registros.map(r=>{
      const a=r.marca==='A',t=r.marca==='T',j=r.marca==='J';
      const av=j?'avj':a?'ava':'avt'; const rb=j?'rbj':a?'rba':'rbt';
      return `<div class="li"><div class="av ${av}">${r.nombre[0]}</div><div><div class="ln">${r.nombre}</div><div class="ls">${j?'✅ Justificado':a?'🟢 Asistió':'🟠 Tardanza'}${j&&r.motivo?' — '+r.motivo:''}</div></div><span class="rb ${rb}">${r.hora}</span></div>`;
    }).join('');
    stats(d.registros.length, d.total);
  }catch(e){}
}

async function cargarFaltantes(){
  try{
    const r=await fetch('/faltantes_hoy',{credentials:'include'});const d=await r.json();
    const el=document.getElementById('lista-falt');
    if(!d.faltantes.length){el.innerHTML='<div class="empty-msg" style="color:#81c784">✅ ¡Todos presentes hoy!</div>';return;}
    el.innerHTML=`<div style="font-size:.72rem;color:#546e7a;margin-bottom:8px">${d.faltantes.length} alumnos sin registrar</div>`+
    d.faltantes.map(n=>`<div class="li"><div class="av avf">${n[0]}</div><div><div class="ln">${n}</div><div class="ls">Sin registrar</div></div><span class="rb rbf">Falta</span></div>`).join('');
  }catch(e){}
}

async function cargarGenero(){
  try{
    const r=await fetch('/registros_hoy',{credentials:'include'});const d=await r.json();
    const ninos=d.registros.filter(p=>p.genero==='nino');
    const ninas=d.registros.filter(p=>p.genero==='nina');
    const el=document.getElementById('gen-content');
    if(!d.registros.length){el.innerHTML='<div class="empty-msg">No hay registros aún</div>';return;}
    let html=`<div class="genero-grid"><div class="genero-col"><div class="genero-titulo titulo-nino">👦 Niños (${ninos.length})</div>`;
    html+=ninos.length?ninos.map(n=>`<div class="alumno-item">👦 ${n.nombre.split(',')[0]}</div>`).join(''):'<div class="empty-msg">Ninguno</div>';
    html+=`</div><div class="genero-col"><div class="genero-titulo titulo-nina">👧 Niñas (${ninas.length})</div>`;
    html+=ninas.length?ninas.map(n=>`<div class="alumno-item">👧 ${n.nombre.split(',')[0]}</div>`).join(''):'<div class="empty-msg">Ninguna</div>';
    html+='</div></div>';
    el.innerHTML=html;
  }catch(e){}
}

async function formarGrupos(){
  const tam=parseInt(document.getElementById('num-grp').value)||4;
  try{
    const r=await fetch('/registros_hoy',{credentials:'include'});const d=await r.json();
    const presentes=d.registros;
    if(!presentes.length){document.getElementById('grp-result').innerHTML='<div class="empty-msg">No hay alumnos presentes</div>';return;}
    const shuffled=[...presentes].sort(()=>Math.random()-.5);
    const grupos=[];
    for(let i=0;i<shuffled.length;i+=tam) grupos.push(shuffled.slice(i,i+tam));
    const cols=['#1565c0','#2e7d32','#b71c1c','#e65100','#4a148c','#00695c'];
    let html='<div class="grupos-container">';
    grupos.forEach((g,i)=>{
      html+=`<div class="grupo-card" style="background:${cols[i%cols.length]}18;border:1px solid ${cols[i%cols.length]}40">`;
      html+=`<div class="grupo-titulo" style="color:${cols[i%cols.length]}">Grupo ${i+1} · ${g.length} integrantes</div>`;
      html+='<div class="grupo-miembros">';
      g.forEach(m=>{html+=`<div class="miembro">${m.genero==='nino'?'👦':'👧'} ${m.nombre.split(',')[0]}</div>`;});
      html+='</div></div>';
    });
    html+='</div>';
    document.getElementById('grp-result').innerHTML=html;
    document.getElementById('btn-regen').style.display='block';
  }catch(e){document.getElementById('grp-result').innerHTML='<div class="empty-msg">Error</div>';}
}

let histInitDone=false;
function initHistorial(){
  if(histInitDone) return; histInitDone=true;
  const sel=document.getElementById('hist-alumno');
  sel.innerHTML='<option value="">-- Selecciona un alumno --</option>';
  ALUMNOS_JS.forEach(a=>{
    const p=a.split(',');const ap=p[0]||'';const nm=p[1]?p[1].trim():'';
    const opt=document.createElement('option');
    opt.value=a; opt.textContent=nm+' '+ap; sel.appendChild(opt);
  });
  const selMes=document.getElementById('hist-mes');
  selMes.innerHTML='<option value="">-- Selecciona un mes --</option>';
  const ano=new Date().getFullYear();
  for(let m=1;m<=12;m++){
    const opt=document.createElement('option');
    opt.value=m; opt.textContent=MESES_NOMBRES[m-1]+' '+ano; selMes.appendChild(opt);
  }
}

function histCambioAlumno(){
  const alumno=document.getElementById('hist-alumno').value;
  const selMes=document.getElementById('hist-mes');
  if(alumno){selMes.style.display='block';document.getElementById('hist-result').innerHTML='<div class="empty-msg">Selecciona un mes</div>';}
  else{selMes.style.display='none';}
  selMes.value='';
}

async function cargarHistorial(){
  const alumno=document.getElementById('hist-alumno').value;
  const mes=document.getElementById('hist-mes').value;
  const ano=new Date().getFullYear();
  if(!alumno||!mes) return;
  try{
    const r=await fetch(`/historial?alumno=${encodeURIComponent(alumno)}&mes=${mes}&ano=${ano}`,{credentials:'include'});
    const d=await r.json();
    const el=document.getElementById('hist-result');
    if(!d.dias||!d.dias.length){el.innerHTML='<div class="empty-msg">Sin datos para este mes</div>';return;}
    const ca=d.dias.filter(x=>x.marca==='A').length,ct=d.dias.filter(x=>x.marca==='T').length;
    const cf=d.dias.filter(x=>x.marca==='F').length,cj=d.dias.filter(x=>x.marca==='J').length;
    let html=`<div class="hist-resumen">
      <span class="hist-badge" style="background:rgba(46,125,50,.2);color:#81c784">A: ${ca}</span>
      <span class="hist-badge" style="background:rgba(230,81,0,.15);color:#ffb74d">T: ${ct}</span>
      <span class="hist-badge" style="background:rgba(183,28,28,.2);color:#ef9a9a">F: ${cf}</span>
      <span class="hist-badge" style="background:rgba(91,107,240,.2);color:#9fa8da">J: ${cj}</span>
    </div><div class="hist-tabla">`;
    d.dias.forEach(x=>{
      const col=x.marca==='A'?'#81c784':x.marca==='T'?'#ffb74d':x.marca==='J'?'#9fa8da':'#ef9a9a';
      const bg=x.marca==='A'?'rgba(46,125,50,.12)':x.marca==='T'?'rgba(230,81,0,.1)':x.marca==='J'?'rgba(91,107,240,.1)':'rgba(183,28,28,.12)';
      html+=`<div class="hist-row" style="background:${bg}">
        <span class="hist-dia">${x.dia}</span>
        <span class="hist-letra" style="color:${col}">${x.marca}</span>
        <span style="font-size:.78rem;color:#cfd8dc">${['','Lun','Mar','Mié','Jue','Vie'][x.dow]||''}</span>
        ${x.marca==='J'&&x.motivo?`<span style="font-size:.7rem;color:#9fa8da;flex:1;padding-left:6px">${x.motivo}</span>`:''}
        <span class="hist-hora">${x.hora}</span>
      </div>`;
    });
    html+='</div>';
    el.innerHTML=html;
  }catch(e){document.getElementById('hist-result').innerHTML='<div class="empty-msg">Error</div>';}
}

async function cargarMeses(){
  try{
    const r=await fetch('/resumen_meses',{credentials:'include'});const d=await r.json();
    const n=new Date();const mc=n.getMonth()+1;
    const g=document.getElementById('meses-grid');
    g.innerHTML=d.meses.map(m=>{
      const ac=(m.mes===mc&&m.ano===new Date().getFullYear());
      const td=m.registros>0;
      return `<div class="mes-card ${ac?'mes-activo':''}">
        <div><div class="mes-nombre">${ac?'📌 ':''}${m.nombre} ${m.ano}</div>
        <div class="mes-regs ${td?'mes-tiene':'mes-vacio'}">${td?'✓ '+m.registros+' registros':'Sin registros'}</div></div>
        <button class="btn-dl-mes" onclick="descargarMes(${m.mes},${m.ano})">⬇ Excel</button>
      </div>`;
    }).join('');
  }catch(e){document.getElementById('meses-grid').innerHTML='<div class="empty-msg">Error</div>';}
}

function descargarMes(mes,ano){window.location.href='/descargar?ano='+ano+'&mes='+mes;}
function dlExcel(){const n=new Date();window.location.href='/descargar?ano='+n.getFullYear()+'&mes='+(n.getMonth()+1);}

// Arrancar
checkAuth();
</script>
</body>
</html>"""

if __name__ == "__main__":
    if DATABASE_URL:
        try:
            init_db()
            print("✅ Base de datos inicializada")
        except Exception as e:
            print(f"⚠️  Error init DB: {e}")
    port = int(os.environ.get("PORT", 7860))
    app.run(host="0.0.0.0", port=port, debug=False)
